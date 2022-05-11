from flask import Flask, request, render_template
from werkzeug.utils import secure_filename
import re
import operator

app = Flask(__name__)

@app.route('/')
def my_form():
    return render_template('form.html')

@app.route('/', methods=['POST'])
def my_form_post():
    
    #convert to lowercase
    text1 = request.form['text1'].lower()
    textx = text1.upper() 
    input_text = request.form['text1']
    import re
    input_text = re.sub("[a-zA-Z0-9]", "", input_text)
    word = input_text.lower()
    import string
    spec_chars = string.punctuation + '\n\xa0«»\t—…[]' 
    word = "".join([ch for ch in word if ch not in spec_chars])
    def remove_chars_from_text(word, chars):
        return "".join([ch for ch in word if ch not in chars])
    word = remove_chars_from_text(word, spec_chars)
    word = remove_chars_from_text(word, string.digits)
    import nltk
    nltk.download('punkt')
    from nltk import word_tokenize
    text_tokens = word_tokenize(word)
    g = 0
    text_tokens_int = []
    while g < len(text_tokens):
        word = text_tokens[g]
        text_tokens_int.append(word)
        g += 1
    ### HERE START PSYCHOSEMANTIC CODE
    text_w = text_tokens_int
    r = 0
    St_list = []
    while r < len(text_w):
        word = text_w[r]        
        import openpyxl
        wb = openpyxl.reader.excel.load_workbook(filename="psy_dataset.xlsx")
        wb.active = 0
        sheet = wb.active
        i = 0
        n = 1
        word_l = list(word.strip(" "))
        soft_let = 'е ё и ь ю я'.split()
        soft_let_pos = 'б в г д з к л м н п р с т ф х'.split()
        result = list(set(word_l) - set(soft_let_pos))
        slpn = 0
        sln = 0
        for i in range(len(word_l)):
            slpn = 0
            while slpn < len(soft_let_pos):
                if word_l[i] == soft_let_pos[slpn]:
                    try:
                        sln = 0
                        while sln < len(soft_let):
                            if word_l[i+1] == soft_let[sln]:               
                                try:
                                    word_l[i] = word_l[i] + '*'
                                    sln += 1
                                    slpn += 1
                                except IndexError:
                                    #print ('nah')
                                    sln += 1
                                    slpn += 1
                            else:
                                #print ('no')
                                sln += 1
                                slpn += 1
                    except IndexError:
                        slpn += 1
                else:
                    slpn += 1
        i = 0
        n = 1
        cr = 1 #row
        cc = 1 #column
        word_lgood = []
        for i in range(len(word_l)):
            while True:
                word_l[i] != sheet.cell(cr,cc).value
                cr += 1
                if word_l[i] == sheet.cell(cr,cc).value:
                    word_l[i] = sheet.cell(cr,cc+1).value
                    word_lgood.append(sheet.cell(cr,cc+4).value)
                    i += 1
                    cr = 1
                    break
        PmaxPi = []
        for i in range(len(word_l)): #Pmax/Pi
            try:
                Pi = max(word_l)/word_l[i]
            except ArithmeticError:
                Pi = 0
            PmaxPi.append(round(Pi,2))
        ki = PmaxPi
        ki[0] = ki[0]*4
        import numpy as np
        xiki = np.array(ki)*np.array(word_lgood)
        St = sum(xiki)/sum(ki)
        St_list.append(round(St, 3))
        r += 1
    St_avrg = sum(St_list)/len(St_list)
    St_avrg = round(St_avrg, 3)
    collect_list = dict(zip(text_w, St_list))
    xyr = sorted(collect_list.items(), key=operator.itemgetter(1))
    xyr_top = []
    h = 0
    while h < 10:
        try:
            xyr_top.append(xyr[h])
            h += 1
        except IndexError:
            xyr_top.append('-')
            h += 1

            
    import numpy as np
    import matplotlib.pyplot as plt


    data = [[ 66386, 174296,  75131, 577908,  32015],
            [ 58230, 381139,  78045,  99308, 160454],
            [ 89135,  80552, 152558, 497981, 603535],
            [ 78415,  81858, 150656, 193263,  69638],
            [139361, 331509, 343164, 781380,  52269]]

    columns = ('Freeze', 'Wind', 'Flood', 'Quake', 'Hail')
    rows = ['%d year' % x for x in (100, 50, 20, 10, 5)]

    values = np.arange(0, 2500, 500)
    value_increment = 1000

    # Get some pastel shades for the colors
    colors = plt.cm.BuPu(np.linspace(0, 0.5, len(rows)))
    n_rows = len(data)

    index = np.arange(len(columns)) + 0.3
    bar_width = 0.4

    # Initialize the vertical-offset for the stacked bar chart.
    y_offset = np.zeros(len(columns))

    # Plot bars and create text labels for the table
    cell_text = []
    for row in range(n_rows):
        plt.bar(index, data[row], bar_width, bottom=y_offset, color=colors[row])
        y_offset = y_offset + data[row]
        cell_text.append(['%1.1f' % (x / 1000.0) for x in y_offset])
    # Reverse colors and text labels to display the last value at the top.
    colors = colors[::-1]
    cell_text.reverse()

    # Add a table at the bottom of the axes
    the_table = plt.table(cellText=cell_text,
                          rowLabels=rows,
                          rowColours=colors,
                          colLabels=columns,
                          loc='bottom')

    # Adjust layout to make room for the table:
    plt.subplots_adjust(left=0.2, bottom=0.2)

    plt.ylabel("Loss in ${0}'s".format(value_increment))
    plt.yticks(values * value_increment, ['%d' % val for val in values])
    plt.xticks([])
    plt.title('Loss by Disaster')
    plt.savefig('saved_figure.png')

    import Image
    myImage = Image.open("saved_figure.png");

    table_x = myImage.show()
    

    return render_template('form.html', final=len(text_tokens),
                           text1=len(text_tokens),
                           text2=len(text_tokens),
                           text3=St_avrg,
                           text4=xyr_top[0],
                           text5=xyr_top[1],
                           text6=xyr_top[2],
                           text7=xyr_top[3],
                           text8=xyr_top[4],
                           text9=xyr_top[5],
                           text10=xyr_top[6],
                           text11=xyr_top[7],
                           text12=xyr_top[8],
                           text13=table_x,
                           text14=table_x)
                            

if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5002, threaded=True)
